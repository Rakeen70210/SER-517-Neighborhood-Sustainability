# This code gets the cities list per state from www.city-data.com
# And store it in data/ state-wise
import time

import requests
import xlsxwriter
from bs4 import BeautifulSoup
import csv
import openpyxl
import os
import glob
import difflib


def get_income_details(link):
    html_text = requests.get(link).text
    soup = BeautifulSoup(html_text, 'html.parser')

    cityDetails = {}
    # primary panels
    for div in soup.find_all(id="tm"):
        key = div.find("div", class_="panel-heading").text
        body = div.find("div", class_="panel-body")
        # print("key (races): ", key)
        cityDetails[key] = {}

        # all text fields (ignoring graphs)
        for well in div.find_all("div", class_="hgraph"):
            fieldAbout = well.find("b").text.lower()

            if "median" in fieldAbout:
                td_ptr = well.find_next("td")
                forCity = td_ptr.text
                td_ptr = td_ptr.find_next("td")
                det1 = td_ptr.text

                td_ptr = td_ptr.find_next("td")
                forState = td_ptr.text
                td_ptr = td_ptr.find_next("td")
                det2 = td_ptr.text
                cityDetails[key][fieldAbout] = [[forCity, det1], [forState, det2]]
    return cityDetails


# def getCitiesMaps(path):
#     citiesMap = {}
#     with open(path, "r") as readObj:
#         csv_reader = csv.reader(readObj)
#         header = next(csv_reader)
#
#         for row in csv_reader:
#             key = row[1]+":"+row[4]+":"+row[2]
#             citiesMap[key] = row
#
#     return [header, citiesMap]

def getAllCitiesMap():
    path = os.getcwd()
    excel_files = glob.glob(os.path.join("../data/", "*.xlsx"))
    cityDataCities = {}

    for file in excel_files:
        wrkbk = openpyxl.load_workbook(file)
        sh = wrkbk.active
        fileName = file.split("\\")[-1]
        stateName = " ".join(fileName.split()[:-1]).lower()

        # for testing 12 cities per state
        # for row in sh.iter_rows(min_row=0, min_col=0, max_row=12, max_col=3):
        for row in sh.iter_rows(min_row=0, min_col=0, max_col=3):
            cityName = row[0].value
            href = row[2].value
            if cityName and href:
                cityName = cityName.lower().replace("-", " ")
                cityDataCities[cityName + ":" + stateName] = incomeBaseURL + href
                # print(cityName+ ":" + stateName)
    return cityDataCities


if __name__ == '__main__':
    #    income_base = "www.city-data.com/income/income-Andover-Florida.html"
    incomeBaseURL = "http://www.city-data.com/income/income-"

    # As details are per races
    incomeCategoryParts = ["all_", "white_", "black_or_african_american_", "asian_",
                           "hispanic_or_latino_", "american_indian_and_alaska_native_", "multirace_",
                           "other_"]
    incomeCategoryPartsKeywords = ["all residents", "white residents", "black or african american residents", "asian residents", "hispanic or latino residents", "american indian and alaska native residents", "multirace residents", "other residents"]

    incomeColumnIndices = [
        "median_household_income_2019",
        "change_in_median_household_income_between_2000_and_2019",
        "median_non-family_income_2019",
        "change_in_median_non-family_income_between_2000_and_2019",
        "median_per_capita_income_2019",
        "change_in_median_per_capita_income_between_2000_and_2019"
    ]
    incomeColumnIndicesKeywords = [
        "median household income",
        "change in median household income",
        "median non-family income",
        "change in median non-family income",
        "median per capita income",
        "change in median per capita income"
    ]

    exceptionsFor500 = {
        "nashville:tennessee": "nashville davidson:tennessee",
        "lexington:kentucky": "lexington fayette:kentucky",
        "san buenaventura:california": 'san buenaventura (ventura):california'
    }

    # states not present on city-data.com
    exceptionsStates = {"puerto rico", "village of islands"}

    # name of csv output file
    outputFile = open("../output/updatedCities.csv", 'w', newline='', encoding='utf-8')
    csvWriter = csv.writer(outputFile)

    def writeCityDetails(row, cityIncomeDetails):
        rowLenOffset = len(row)

        # initialize all the values
        row += ["N"] * len(incomeCategoryParts) * len(incomeColumnIndices)

        def getIndexOfColumn(pre, suff):
            for i, prefKwrd in enumerate(incomeCategoryPartsKeywords):
                if prefKwrd in pre:
                    for j, sufKwrd in enumerate(incomeColumnIndicesKeywords):
                        if sufKwrd in suff:
                            # common suffix handler
                            if not j % 2 and "change in" in suff:
                                return len(incomeColumnIndices) * i + j + 1
                            return len(incomeColumnIndices) * i + j
            return -1

        for pre in cityIncomeDetails.keys():
            for suf, val in cityIncomeDetails[pre].items():
                index = getIndexOfColumn(pre.lower(), suf.lower())
                if(index != -1):
                    row[rowLenOffset + index] = val[0][1]
        print("Writing details for: ", row[:5])
        csvWriter.writerow(row)

    # For Manual-review
    toCheck = []
    totalPassed = 0
    totalFailed = 0
    targetFailed = 0
    # generate city-data cities map
    cityDataCities = getAllCitiesMap()
    cityDataCitiesList = cityDataCities.keys()


    def findClosestMatch(key):
        return difflib.get_close_matches(key, cityDataCitiesList, cutoff=0.8, n=3)


    with open("../finalAllCities.csv", "r") as readObj:
        csvReader = csv.reader(readObj)
        headers = next(csvReader)
        baseHeadersLen = len(headers)


        # add new columns
        def addNewColumns():
            for pre in incomeCategoryParts:
                for suff in incomeColumnIndices:
                    headers.append(pre + suff)


        addNewColumns()
        # writing the headers
        csvWriter.writerow(headers)
        # print("headers: ", headers)

        for row in csvReader:
            cityName = row[1].lower().replace("-", " ")
            stateName = row[2].lower()
            key = cityName + ":" + stateName
            cityIncomeDetails = {}

            if key in cityDataCities:
                cityIncomeDetails = get_income_details(cityDataCities[key])
                totalPassed += 1
                time.sleep(10)
            elif key in exceptionsFor500:
                cityIncomeDetails = get_income_details(cityDataCities[exceptionsFor500[key]])
                totalPassed += 1
                time.sleep(10)
            else:
                if stateName not in exceptionsStates:
                    # closestMatchList = findClosestMatch(key)
                    # print("closest match for ", key, " are ", closestMatchList)
                    totalFailed += 1
            writeCityDetails(row, cityIncomeDetails)
    # todo
    # Target cities score (check with 500 cities)

    # print("toCheck ", toCheck)
    print("totalCities= ", len(cityDataCities))
    print("totalPassed= ", totalPassed, " and totalFailed= ", totalFailed)
